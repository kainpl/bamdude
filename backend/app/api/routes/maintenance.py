"""Maintenance tracking API routes."""

import logging
import math
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from backend.app.core.auth import RequirePermission
from backend.app.core.database import get_db
from backend.app.core.permissions import Permission
from backend.app.models.maintenance import MaintenanceHistory, MaintenanceType, PrinterMaintenance
from backend.app.models.printer import Printer
from backend.app.models.user import User
from backend.app.schemas.maintenance import (
    MaintenanceHistoryResponse,
    MaintenanceStatus,
    MaintenanceTypeCreate,
    MaintenanceTypeResponse,
    MaintenanceTypeUpdate,
    PerformMaintenanceRequest,
    PrinterMaintenanceOverview,
    PrinterMaintenanceResponse,
    PrinterMaintenanceUpdate,
)
from backend.app.services.notification_service import notification_service
from backend.app.utils.printer_models import get_rod_type

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/maintenance", tags=["maintenance"])

# Default maintenance types
DEFAULT_MAINTENANCE_TYPES = [
    # Carbon rod models only (X1/P1)
    {
        "type_code": "clean_carbon_rods",
        "name": "Clean Carbon Rods",
        "description": "Wipe carbon rods with a dry cloth",
        "default_interval_hours": 100.0,
        "icon": "Sparkles",
        "printer_models": '["X1C", "X1", "X1E", "P1P", "P1S"]',
    },
    # Steel rod models only (P2S)
    {
        "type_code": "lubricate_steel_rods",
        "name": "Lubricate Steel Rods",
        "description": "Apply lubricant to steel rods for smooth motion",
        "default_interval_hours": 50.0,
        "icon": "Droplet",
        "printer_models": '["P2S"]',
    },
    {
        "type_code": "clean_steel_rods",
        "name": "Clean Steel Rods",
        "description": "Wipe steel rods with a dry cloth",
        "default_interval_hours": 100.0,
        "icon": "Sparkles",
        "printer_models": '["P2S"]',
    },
    # Linear rail models only (A1/H2)
    {
        "type_code": "lubricate_linear_rails",
        "name": "Lubricate Linear Rails",
        "description": "Apply lubricant to linear rails for smooth motion",
        "default_interval_hours": 50.0,
        "icon": "Droplet",
        "printer_models": '["A1", "A1 Mini", "H2D", "H2D Pro", "H2C", "H2S"]',
    },
    {
        "type_code": "clean_linear_rails",
        "name": "Clean Linear Rails",
        "description": "Wipe linear rails with a dry cloth to remove dust and debris",
        "default_interval_hours": 100.0,
        "icon": "Sparkles",
        "printer_models": '["A1", "A1 Mini", "H2D", "H2D Pro", "H2C", "H2S"]',
    },
    # Universal (all models)
    {
        "type_code": "clean_nozzle",
        "name": "Clean Nozzle/Hotend",
        "description": "Clean nozzle exterior and perform cold pull if needed",
        "default_interval_hours": 100.0,
        "icon": "Flame",
        "printer_models": '["*"]',
    },
    {
        "type_code": "check_belt_tension",
        "name": "Check Belt Tension",
        "description": "Verify and adjust belt tension for X/Y axes",
        "default_interval_hours": 200.0,
        "icon": "Ruler",
        "printer_models": '["*"]',
    },
    {
        "type_code": "clean_build_plate",
        "name": "Clean Build Plate",
        "description": "Deep clean build plate with IPA or soap",
        "default_interval_hours": 25.0,
        "icon": "Square",
        "printer_models": '["*"]',
    },
    {
        "type_code": "check_ptfe_tube",
        "name": "Check PTFE Tube",
        "description": "Inspect PTFE tube for wear or discoloration",
        "default_interval_hours": 500.0,
        "icon": "Cable",
        "printer_models": '["*"]',
    },
]

# System types that only apply to printers with a specific rod/rail type.
# Keyed by type_code. Types not listed here apply to all printers.
_ROD_TYPE_REQUIREMENTS: dict[str, str] = {
    "clean_carbon_rods": "carbon",
    "lubricate_steel_rods": "steel_rod",
    "clean_steel_rods": "steel_rod",
    "lubricate_linear_rails": "linear_rail",
    "clean_linear_rails": "linear_rail",
}


def _should_apply_to_printer(type_code: str | None, printer_model: str | None) -> bool:
    """Check if a system maintenance type should apply to a given printer model."""
    if not type_code:
        return True
    rod_requirement = _ROD_TYPE_REQUIREMENTS.get(type_code)
    if rod_requirement is None:
        return True  # Not model-specific, applies to all

    rod_type = get_rod_type(printer_model)
    if rod_type is None:
        # Unknown model — default to carbon rods (legacy behavior)
        return rod_requirement == "carbon"

    return rod_type == rod_requirement


async def get_printer_total_hours(db: AsyncSession, printer_id: int) -> float:
    """Calculate total active hours for a printer from runtime counter plus offset.

    Uses the runtime_seconds counter which tracks actual machine active time
    (RUNNING and PAUSE states), including calibration, heating, and printing.
    """
    # Get printer runtime and offset
    result = await db.execute(
        select(Printer.runtime_seconds, Printer.print_hours_offset).where(Printer.id == printer_id)
    )
    row = result.one_or_none()
    if not row:
        return 0.0

    runtime_seconds = row[0] or 0
    offset = row[1] or 0.0

    runtime_hours = runtime_seconds / 3600.0
    return runtime_hours + offset


async def ensure_default_types(db: AsyncSession) -> None:
    """Ensure default maintenance types exist. Never deletes — only adds missing ones."""
    result = await db.execute(select(MaintenanceType).where(MaintenanceType.is_system.is_(True)))
    existing = result.scalars().all()
    existing_codes = {t.type_code for t in existing if t.type_code}

    added = False
    for type_def in DEFAULT_MAINTENANCE_TYPES:
        if type_def["type_code"] in existing_codes:
            continue
        new_type = MaintenanceType(
            type_code=type_def["type_code"],
            name=type_def["name"],
            description=type_def["description"],
            default_interval_hours=type_def["default_interval_hours"],
            icon=type_def["icon"],
            printer_models=type_def.get("printer_models", '["*"]'),
            is_system=True,
        )
        db.add(new_type)
        added = True

    if added:
        await db.commit()


# ============== Maintenance Types ==============


@router.get("/types", response_model=list[MaintenanceTypeResponse])
async def get_maintenance_types(
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get all maintenance types."""
    await ensure_default_types(db)
    result = await db.execute(
        select(MaintenanceType)
        .where(MaintenanceType.is_deleted.is_(False))
        .order_by(MaintenanceType.is_system.desc(), MaintenanceType.name)
    )
    return result.scalars().all()


@router.post("/types", response_model=MaintenanceTypeResponse)
async def create_maintenance_type(
    data: MaintenanceTypeCreate,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_CREATE),
):
    """Create a custom maintenance type."""
    import json as _json

    new_type = MaintenanceType(
        name=data.name,
        description=data.description,
        default_interval_hours=data.default_interval_hours,
        interval_type=data.interval_type,
        icon=data.icon,
        printer_models=_json.dumps(data.printer_models),
        is_system=False,
    )
    db.add(new_type)
    await db.commit()
    await db.refresh(new_type)

    # Set type_code for custom types after id is assigned
    new_type.type_code = f"custom_{new_type.id}"
    await db.commit()
    await db.refresh(new_type)
    return new_type


@router.patch("/types/{type_id}", response_model=MaintenanceTypeResponse)
async def update_maintenance_type(
    type_id: int,
    data: MaintenanceTypeUpdate,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_UPDATE),
):
    """Update a maintenance type."""
    result = await db.execute(select(MaintenanceType).where(MaintenanceType.id == type_id))
    maint_type = result.scalar_one_or_none()
    if not maint_type:
        raise HTTPException(status_code=404, detail="Maintenance type not found")

    import json as _json

    update_data = data.model_dump(exclude_unset=True)
    if "printer_models" in update_data:
        update_data["printer_models"] = _json.dumps(update_data["printer_models"])
    for key, value in update_data.items():
        setattr(maint_type, key, value)

    await db.commit()
    await db.refresh(maint_type)
    return maint_type


@router.delete("/types/{type_id}")
async def delete_maintenance_type(
    type_id: int,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_DELETE),
):
    """Delete a maintenance type."""
    result = await db.execute(select(MaintenanceType).where(MaintenanceType.id == type_id))
    maint_type = result.scalar_one_or_none()
    if not maint_type:
        raise HTTPException(status_code=404, detail="Maintenance type not found")

    if maint_type.is_system:
        maint_type.is_deleted = True
        await db.commit()
        return {"status": "deleted"}

    await db.delete(maint_type)
    await db.commit()
    return {"status": "deleted"}


@router.post("/types/restore-defaults")
async def restore_default_maintenance_types(
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_DELETE),
):
    """Restore deleted default maintenance types."""
    await ensure_default_types(db)
    result = await db.execute(
        select(MaintenanceType).where(MaintenanceType.is_system.is_(True)).where(MaintenanceType.is_deleted.is_(True))
    )
    deleted_types = result.scalars().all()
    for maint_type in deleted_types:
        maint_type.is_deleted = False

    await db.commit()
    return {"restored": len(deleted_types)}


# ============== Printer Maintenance ==============


async def _get_printer_maintenance_internal(
    printer_id: int,
    db: AsyncSession,
    commit: bool = True,
) -> PrinterMaintenanceOverview:
    """Internal helper to get maintenance overview for a specific printer."""
    await ensure_default_types(db)

    # Get printer
    result = await db.execute(select(Printer).where(Printer.id == printer_id))
    printer = result.scalar_one_or_none()
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    total_hours = await get_printer_total_hours(db, printer_id)

    # Get all maintenance types
    result = await db.execute(select(MaintenanceType).where(MaintenanceType.is_deleted.is_(False)))
    all_types = result.scalars().all()

    # Get printer's maintenance items
    result = await db.execute(
        select(PrinterMaintenance)
        .where(PrinterMaintenance.printer_id == printer_id)
        .options(selectinload(PrinterMaintenance.maintenance_type))
    )
    existing_items = {item.maintenance_type_id: item for item in result.scalars().all()}

    maintenance_items = []
    due_count = 0
    warning_count = 0

    now = datetime.now(timezone.utc)

    for maint_type in all_types:
        # Skip system types that don't apply to this printer model
        # (e.g., "Clean Carbon Rods" for H2D which has steel rods)
        if maint_type.is_system and not _should_apply_to_printer(maint_type.type_code, printer.model):
            continue

        item = existing_items.get(maint_type.id)
        default_interval_type = getattr(maint_type, "interval_type", "hours") or "hours"

        if item:
            interval = item.custom_interval_hours or maint_type.default_interval_hours
            # Use custom interval type if set, otherwise use type's default
            interval_type = getattr(item, "custom_interval_type", None) or default_interval_type
            enabled = item.enabled
            last_performed_hours = item.last_performed_hours
            last_performed_at = item.last_performed_at
            item_id = item.id
        else:
            # Only auto-create maintenance items for system types
            # Custom types need to be manually assigned per printer
            if not maint_type.is_system:
                continue

            # Check if this type applies to this printer's model
            import json as _json

            try:
                type_models = (
                    _json.loads(maint_type.printer_models)
                    if isinstance(maint_type.printer_models, str)
                    else (maint_type.printer_models or ["*"])
                )
            except (ValueError, TypeError):
                type_models = ["*"]
            if "*" not in type_models:
                printer_result = await db.execute(select(Printer.model).where(Printer.id == printer_id))
                printer_model = printer_result.scalar()
                if printer_model and printer_model not in type_models:
                    continue

            # Create default entry for this printer/type
            item = PrinterMaintenance(
                printer_id=printer_id,
                maintenance_type_id=maint_type.id,
                enabled=True,
                last_performed_hours=0.0,
            )
            db.add(item)
            await db.flush()

            interval = maint_type.default_interval_hours
            interval_type = default_interval_type
            enabled = True
            last_performed_hours = 0.0
            last_performed_at = None
            item_id = item.id

        # Calculate status based on interval type
        if interval_type == "days":
            # Time-based: calculate days since last performed
            if last_performed_at:
                # DB stores naive datetimes; treat as UTC for comparison
                if last_performed_at.tzinfo is None:
                    last_performed_at = last_performed_at.replace(tzinfo=timezone.utc)
                days_since = (now - last_performed_at).total_seconds() / 86400.0
            else:
                # Never performed - consider it due
                days_since = interval + 1

            days_until = interval - days_since
            is_due = days_until <= 0
            is_warning = days_until <= (interval * 0.1) and not is_due

            # For compatibility, also set hours values (but they won't be primary)
            hours_since = total_hours - last_performed_hours
            hours_until = 0  # Not applicable for time-based
        else:
            # Print-hours based (default)
            hours_since = total_hours - last_performed_hours
            hours_until = interval - hours_since
            is_due = hours_until <= 0
            is_warning = hours_until <= (interval * 0.1) and not is_due

            # Calculate days for reference
            if last_performed_at:
                if last_performed_at.tzinfo is None:
                    last_performed_at = last_performed_at.replace(tzinfo=timezone.utc)
                days_since = (now - last_performed_at).total_seconds() / 86400.0
            else:
                days_since = None
            days_until = None

        if enabled:
            if is_due:
                due_count += 1
            elif is_warning:
                warning_count += 1

        maintenance_items.append(
            MaintenanceStatus(
                id=item_id,
                printer_id=printer_id,
                printer_name=printer.name,
                printer_model=printer.model,
                maintenance_type_id=maint_type.id,
                maintenance_type_name=maint_type.name,
                maintenance_type_code=maint_type.type_code,
                maintenance_type_icon=maint_type.icon,
                maintenance_type_wiki_url=getattr(maint_type, "wiki_url", None),
                enabled=enabled,
                interval_hours=interval,
                interval_type=interval_type,
                current_hours=total_hours,
                hours_since_maintenance=hours_since,
                hours_until_due=hours_until,
                days_since_maintenance=days_since if interval_type == "days" else None,
                days_until_due=days_until if interval_type == "days" else None,
                is_due=is_due,
                is_warning=is_warning,
                last_performed_at=last_performed_at,
            )
        )

    if commit:
        await db.commit()

    return PrinterMaintenanceOverview(
        printer_id=printer_id,
        printer_name=printer.name,
        printer_model=printer.model,
        total_print_hours=total_hours,
        maintenance_items=maintenance_items,
        due_count=due_count,
        warning_count=warning_count,
    )


@router.get("/printers/{printer_id}", response_model=PrinterMaintenanceOverview)
async def get_printer_maintenance(
    printer_id: int,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get maintenance overview for a specific printer."""
    return await _get_printer_maintenance_internal(printer_id, db, commit=True)


@router.get("/overview", response_model=list[PrinterMaintenanceOverview])
async def get_all_maintenance_overview(
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get maintenance overview for all active printers."""
    await ensure_default_types(db)

    result = await db.execute(select(Printer).where(Printer.is_active.is_(True)))
    printers = result.scalars().all()

    overviews = []
    for printer in printers:
        # Don't commit after each printer, commit once at the end
        overview = await _get_printer_maintenance_internal(printer.id, db, commit=False)
        overviews.append(overview)

    # Commit any new maintenance items created
    await db.commit()

    return overviews


@router.patch("/items/{item_id}", response_model=PrinterMaintenanceResponse)
async def update_printer_maintenance(
    item_id: int,
    data: PrinterMaintenanceUpdate,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_UPDATE),
):
    """Update a printer maintenance item (e.g., custom interval, enabled)."""
    result = await db.execute(
        select(PrinterMaintenance)
        .where(PrinterMaintenance.id == item_id)
        .options(selectinload(PrinterMaintenance.maintenance_type))
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Maintenance item not found")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(item, key, value)

    await db.commit()
    await db.refresh(item)
    return item


@router.post("/printers/{printer_id}/assign/{type_id}", response_model=PrinterMaintenanceResponse)
async def assign_maintenance_type(
    printer_id: int,
    type_id: int,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_CREATE),
):
    """Assign a maintenance type to a specific printer (for custom types)."""
    # Verify printer exists
    result = await db.execute(select(Printer).where(Printer.id == printer_id))
    printer = result.scalar_one_or_none()
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    # Verify maintenance type exists
    result = await db.execute(select(MaintenanceType).where(MaintenanceType.id == type_id))
    maint_type = result.scalar_one_or_none()
    if not maint_type:
        raise HTTPException(status_code=404, detail="Maintenance type not found")

    # Check if already assigned
    result = await db.execute(
        select(PrinterMaintenance).where(
            PrinterMaintenance.printer_id == printer_id,
            PrinterMaintenance.maintenance_type_id == type_id,
        )
    )
    existing = result.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Maintenance type already assigned to this printer")

    # Create the assignment
    item = PrinterMaintenance(
        printer_id=printer_id,
        maintenance_type_id=type_id,
        enabled=True,
        last_performed_hours=0.0,
    )
    db.add(item)
    await db.commit()

    # Re-fetch with relationship loaded for response serialization
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(PrinterMaintenance)
        .options(selectinload(PrinterMaintenance.maintenance_type))
        .where(PrinterMaintenance.id == item.id)
    )
    item = result.scalar_one()

    return item


@router.delete("/items/{item_id}")
async def remove_maintenance_item(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_DELETE),
):
    """Remove a maintenance item (unassign a custom type from a printer)."""
    result = await db.execute(
        select(PrinterMaintenance)
        .where(PrinterMaintenance.id == item_id)
        .options(selectinload(PrinterMaintenance.maintenance_type))
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Maintenance item not found")

    # Only allow removing custom (non-system) types
    if item.maintenance_type.is_system:
        raise HTTPException(status_code=400, detail="Cannot remove system maintenance types")

    await db.delete(item)
    await db.commit()

    return {"status": "removed"}


@router.post("/items/{item_id}/perform", response_model=MaintenanceStatus)
async def perform_maintenance(
    item_id: int,
    data: PerformMaintenanceRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User | None = RequirePermission(Permission.MAINTENANCE_UPDATE),
):
    """Mark maintenance as performed (reset the counter)."""
    result = await db.execute(
        select(PrinterMaintenance)
        .where(PrinterMaintenance.id == item_id)
        .options(selectinload(PrinterMaintenance.maintenance_type))
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Maintenance item not found")

    # Get printer for name
    result = await db.execute(select(Printer).where(Printer.id == item.printer_id))
    printer = result.scalar_one()

    # Get current hours
    current_hours = await get_printer_total_hours(db, item.printer_id)

    # Create history entry
    history = MaintenanceHistory(
        printer_maintenance_id=item.id,
        hours_at_maintenance=current_hours,
        notes=data.notes,
        performed_by_user_id=current_user.id if current_user else None,
    )
    db.add(history)

    # Update item
    item.last_performed_at = datetime.now(timezone.utc)
    item.last_performed_hours = current_hours

    await db.commit()

    # MQTT relay - publish maintenance reset
    try:
        from backend.app.services.mqtt_relay import mqtt_relay

        await mqtt_relay.on_maintenance_reset(
            printer_id=item.printer_id,
            printer_name=printer.name,
            maintenance_type=item.maintenance_type.name,
        )
    except Exception:
        pass  # Don't fail if MQTT fails

    # Calculate status
    interval = item.custom_interval_hours or item.maintenance_type.default_interval_hours
    interval_type = getattr(item.maintenance_type, "interval_type", "hours") or "hours"
    hours_since = current_hours - item.last_performed_hours
    hours_until = interval - hours_since

    return MaintenanceStatus(
        id=item.id,
        printer_id=item.printer_id,
        printer_name=printer.name,
        printer_model=printer.model,
        maintenance_type_id=item.maintenance_type_id,
        maintenance_type_name=item.maintenance_type.name,
        maintenance_type_code=item.maintenance_type.type_code,
        maintenance_type_icon=item.maintenance_type.icon,
        maintenance_type_wiki_url=getattr(item.maintenance_type, "wiki_url", None),
        enabled=item.enabled,
        interval_hours=interval,
        interval_type=interval_type,
        current_hours=current_hours,
        hours_since_maintenance=hours_since,
        hours_until_due=hours_until if interval_type == "hours" else 0,
        days_since_maintenance=0 if interval_type == "days" else None,
        days_until_due=interval if interval_type == "days" else None,
        is_due=False,
        is_warning=False,
        last_performed_at=item.last_performed_at,
    )


@router.get("/history/export")
async def export_maintenance_history(
    printer_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Export maintenance history as Excel (.xlsx), optionally filtered by printer."""
    import io
    from datetime import datetime as dt

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from sqlalchemy.orm import selectinload

    query = (
        select(MaintenanceHistory)
        .options(
            selectinload(MaintenanceHistory.performed_by_user),
            selectinload(MaintenanceHistory.performed_by_chat),
            selectinload(MaintenanceHistory.printer_maintenance).selectinload(PrinterMaintenance.maintenance_type),
            selectinload(MaintenanceHistory.printer_maintenance).selectinload(PrinterMaintenance.printer),
        )
        .order_by(MaintenanceHistory.performed_at.desc())
    )
    if printer_id is not None:
        query = query.join(
            PrinterMaintenance, MaintenanceHistory.printer_maintenance_id == PrinterMaintenance.id
        ).where(PrinterMaintenance.printer_id == printer_id)
    result = await db.execute(query)
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance History"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00AE42", end_color="00AE42", fill_type="solid")

    headers = ["Date", "Printer", "Type", "Hours", "Performed by", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, h in enumerate(items, 2):
        printer_name = (
            h.printer_maintenance.printer.name if h.printer_maintenance and h.printer_maintenance.printer else ""
        )
        type_name = (
            h.printer_maintenance.maintenance_type.name
            if h.printer_maintenance and h.printer_maintenance.maintenance_type
            else ""
        )
        performed_by = (
            h.performed_by_user.username
            if h.performed_by_user
            else (h.performed_by_chat.label if h.performed_by_chat else "")
        )

        ws.cell(row=row_idx, column=1, value=h.performed_at)
        ws.cell(row=row_idx, column=2, value=printer_name)
        ws.cell(row=row_idx, column=3, value=type_name)
        ws.cell(row=row_idx, column=4, value=round(h.hours_at_maintenance, 1))
        ws.cell(row=row_idx, column=5, value=performed_by)
        ws.cell(row=row_idx, column=6, value=h.notes or "")

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Date column format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "YYYY-MM-DD HH:MM"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"maintenance_history_{dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/history")
async def get_all_maintenance_history(
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=25, ge=1, le=100),
    sort_by: str = Query(default="date"),
    sort_dir: str = Query(default="desc"),
    printer_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get all maintenance history, optionally filtered by printer."""
    from backend.app.models.user import User as UserModel

    # Count with optional printer filter
    count_query = select(func.count()).select_from(MaintenanceHistory)
    if printer_id is not None:
        count_query = count_query.join(PrinterMaintenance).where(PrinterMaintenance.printer_id == printer_id)
    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0

    # Build order clause
    if sort_by == "user":
        order_col = UserModel.username
        query = select(MaintenanceHistory).outerjoin(UserModel, MaintenanceHistory.performed_by_user_id == UserModel.id)
    else:
        order_col = MaintenanceHistory.performed_at
        query = select(MaintenanceHistory)

    order = order_col.asc() if sort_dir == "asc" else order_col.desc()

    # Apply printer filter
    if printer_id is not None:
        if sort_by != "user":  # user sort already has a join
            query = query.join(PrinterMaintenance, MaintenanceHistory.printer_maintenance_id == PrinterMaintenance.id)
        else:
            query = query.join(PrinterMaintenance, MaintenanceHistory.printer_maintenance_id == PrinterMaintenance.id)
        query = query.where(PrinterMaintenance.printer_id == printer_id)

    result = await db.execute(
        query.options(
            selectinload(MaintenanceHistory.performed_by_user),
            selectinload(MaintenanceHistory.performed_by_chat),
            selectinload(MaintenanceHistory.printer_maintenance).selectinload(PrinterMaintenance.maintenance_type),
            selectinload(MaintenanceHistory.printer_maintenance).selectinload(PrinterMaintenance.printer),
        )
        .order_by(order)
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    items = result.scalars().all()

    return {
        "items": [
            {
                "id": h.id,
                "performed_at": h.performed_at.isoformat() if h.performed_at else None,
                "hours_at_maintenance": h.hours_at_maintenance,
                "notes": h.notes,
                "printer_name": h.printer_maintenance.printer.name
                if h.printer_maintenance and h.printer_maintenance.printer
                else None,
                "maintenance_type_name": h.printer_maintenance.maintenance_type.name
                if h.printer_maintenance and h.printer_maintenance.maintenance_type
                else None,
                "performed_by_user_id": h.performed_by_user_id,
                "performed_by_username": h.performed_by_user.username if h.performed_by_user else None,
                "performed_by_chat_id": h.performed_by_chat_id,
                "performed_by_chat_label": h.performed_by_chat.label if h.performed_by_chat else None,
            }
            for h in items
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "last_page": math.ceil(total / per_page) if total > 0 else 1,
    }


@router.get("/items/{item_id}/history", response_model=list[MaintenanceHistoryResponse])
async def get_maintenance_history(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get maintenance history for a specific item."""
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(MaintenanceHistory)
        .options(
            selectinload(MaintenanceHistory.performed_by_user),
            selectinload(MaintenanceHistory.performed_by_chat),
        )
        .where(MaintenanceHistory.printer_maintenance_id == item_id)
        .order_by(MaintenanceHistory.performed_at.desc())
    )
    items = result.scalars().all()

    return [
        MaintenanceHistoryResponse(
            id=h.id,
            printer_maintenance_id=h.printer_maintenance_id,
            performed_at=h.performed_at,
            hours_at_maintenance=h.hours_at_maintenance,
            notes=h.notes,
            performed_by_user_id=h.performed_by_user_id,
            performed_by_username=h.performed_by_user.username if h.performed_by_user else None,
            performed_by_chat_id=h.performed_by_chat_id,
            performed_by_chat_label=h.performed_by_chat.label if h.performed_by_chat else None,
        )
        for h in items
    ]


@router.get("/summary")
async def get_maintenance_summary(
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_READ),
):
    """Get a summary of maintenance status across all printers."""
    await ensure_default_types(db)

    result = await db.execute(select(Printer).where(Printer.is_active.is_(True)))
    printers = result.scalars().all()

    total_due = 0
    total_warning = 0
    printers_with_issues = []

    for printer in printers:
        overview = await get_printer_maintenance(printer.id, db)
        total_due += overview.due_count
        total_warning += overview.warning_count
        if overview.due_count > 0 or overview.warning_count > 0:
            printers_with_issues.append(
                {
                    "printer_id": printer.id,
                    "printer_name": printer.name,
                    "due_count": overview.due_count,
                    "warning_count": overview.warning_count,
                }
            )

    return {
        "total_due": total_due,
        "total_warning": total_warning,
        "printers_with_issues": printers_with_issues,
    }


@router.patch("/printers/{printer_id}/hours")
async def set_printer_hours(
    printer_id: int,
    total_hours: float,
    db: AsyncSession = Depends(get_db),
    _: User | None = RequirePermission(Permission.MAINTENANCE_UPDATE),
):
    """Set the total print hours for a printer (adjusts offset to match).

    The offset is calculated as: offset = total_hours - runtime_hours
    Where runtime_hours comes from the runtime_seconds counter that tracks
    actual machine active time (RUNNING/PAUSE states).
    """
    # Get printer
    result = await db.execute(select(Printer).where(Printer.id == printer_id))
    printer = result.scalar_one_or_none()
    if not printer:
        raise HTTPException(status_code=404, detail="Printer not found")

    # Get current runtime hours
    runtime_hours = (printer.runtime_seconds or 0) / 3600.0

    # Calculate needed offset
    printer.print_hours_offset = max(0, total_hours - runtime_hours)

    await db.commit()

    # Check for maintenance items that need attention and send notification
    try:
        await ensure_default_types(db)
        overview = await _get_printer_maintenance_internal(printer_id, db, commit=True)

        items_needing_attention = [
            {
                "id": item.id,
                "name": item.maintenance_type_name,
                "is_due": item.is_due,
                "is_warning": item.is_warning,
            }
            for item in overview.maintenance_items
            if item.enabled and (item.is_due or item.is_warning)
        ]

        if items_needing_attention:
            await notification_service.on_maintenance_due(printer_id, printer.name, items_needing_attention, db)
            logger.info(
                f"Sent maintenance notification for printer {printer_id}: "
                f"{len(items_needing_attention)} items need attention"
            )
    except Exception as e:
        logger.warning("Failed to send maintenance notification: %s", e)

    return {
        "printer_id": printer_id,
        "total_hours": total_hours,
        "runtime_hours": runtime_hours,
        "offset_hours": printer.print_hours_offset,
    }
